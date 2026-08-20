"""Excel çıktısı: data/kampanyalar.xlsx

Yapı:
  "Tüm Kampanyalar"  -> ana tablo, her satır bir kampanya
  "Özet"             -> kategori/banka bazında sayımlar
  <Kategori adı>     -> her kategori için ayrı sekme

Davranış:
  * Excel her koşuda state.json'dan YENİDEN üretilir. Bu kasıtlı: state.json zaten
    tüm geçmişi tutuyor, dolayısıyla tek doğruluk kaynağı odur. Böylece Excel bozulsa
    veya silinse bile bir sonraki koşuda eksiksiz geri gelir.
  * "İlk Görülme" sütunu kampanyanın hangi gün bulunduğunu gösterir; yeni satırlar
    en üstte olacak şekilde sıralanır ve yeşil "YENİ" etiketi alır.
"""
from __future__ import annotations

import logging
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from .categorize import categorize, clean_title, is_ended

log = logging.getLogger(__name__)

FONT = "Arial"
HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(name=FONT, bold=True, color="FFFFFF", size=11)
CELL_FONT = Font(name=FONT, size=10)
LINK_FONT = Font(name=FONT, size=10, color="0563C1", underline="single")
NEW_FILL = PatternFill("solid", fgColor="E2EFDA")
ENDED_FONT = Font(name=FONT, size=10, color="9C9C9C", italic=True)
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

COLUMNS = [
    ("İlk Görülme", 13),
    ("Durum", 9),
    ("Banka", 22),
    ("Kategori", 24),
    ("Etiketler", 30),
    ("Kampanya Başlığı", 70),
    ("Link", 55),
]


def _safe_sheet_name(name: str) -> str:
    """Excel sekme adı: 31 karakter, bu karakterler yasak: : \\ / ? * [ ]"""
    for ch in ":\\/?*[]":
        name = name.replace(ch, "-")
    return name[:31] or "Kategori"


def build_rows(state_seen: dict, use_llm: bool = True) -> list[dict]:
    """state.json'daki 'seen' sözlüğünü Excel satırlarına çevirir."""
    rows: list[dict] = []
    for record in state_seen.values():
        title = (record.get("title") or "").strip()
        url = record.get("url") or ""
        if not title or not url:
            continue
        title = clean_title(title, url)
        if not title:
            continue  # liste/kategori sayfası, kampanya değil

        category, tags = categorize(title, url, use_llm=use_llm)
        first_seen = (record.get("first_seen") or "")[:10]
        rows.append({
            "İlk Görülme": first_seen,
            "Durum": "Bitti" if is_ended(title) else "Aktif",
            "Banka": record.get("source", ""),
            "Kategori": category,
            "Etiketler": ", ".join(dict.fromkeys(tags)),
            "Kampanya Başlığı": title[:300],
            "Link": url,
            "_ended": is_ended(title),
        })

    # En yeni en üstte; aynı gün içinde banka ve kategoriye göre düzenli
    rows.sort(key=lambda r: (r["İlk Görülme"], r["Banka"], r["Kategori"]), reverse=True)
    return rows


def _write_sheet(ws, rows: list[dict], newest_date: str, table_name: str) -> None:
    ws.freeze_panes = "A2"

    for idx, (header, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(vertical="center", horizontal="left")
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(idx)].width = width
    ws.row_dimensions[1].height = 22

    for r_idx, row in enumerate(rows, start=2):
        is_new = row["İlk Görülme"] == newest_date
        for c_idx, (header, _) in enumerate(COLUMNS, start=1):
            value = row[header]
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.border = BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=(header == "Kampanya Başlığı"))
            if header == "Link":
                cell.hyperlink = value
                cell.font = LINK_FONT
            elif row["_ended"]:
                cell.font = ENDED_FONT
            else:
                cell.font = CELL_FONT
            if is_new:
                cell.fill = NEW_FILL

    if rows:
        ref = f"A1:{get_column_letter(len(COLUMNS))}{len(rows) + 1}"
        table = Table(displayName=table_name, ref=ref)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleLight1", showRowStripes=False, showColumnStripes=False
        )
        try:
            ws.add_table(table)   # otomatik filtre + sıralama başlıkları
        except ValueError:
            ws.auto_filter.ref = ref


def _write_summary(ws, rows: list[dict], newest_date: str) -> None:
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 12

    title = ws.cell(row=1, column=1, value="Kampanya Radar — Özet")
    title.font = Font(name=FONT, bold=True, size=14)
    ws.cell(row=2, column=1,
            value=f"Güncelleme: {datetime.now(timezone.utc).strftime('%d.%m.%Y %H:%M')} UTC"
            ).font = Font(name=FONT, size=9, color="808080")
    ws.cell(row=3, column=1,
            value=f"Toplam kampanya: {len(rows)}  |  Bugün eklenen: "
                  f"{sum(1 for r in rows if r['İlk Görülme'] == newest_date)}"
            ).font = Font(name=FONT, size=10)

    def block(start_row: int, heading: str, key: str) -> int:
        ws.cell(row=start_row, column=1, value=heading).font = Font(
            name=FONT, bold=True, size=11)
        for col, head in enumerate(["", "Toplam", "Aktif", "Yeni"], start=1):
            if head:
                c = ws.cell(row=start_row, column=col, value=head)
                c.font = HEADER_FONT
                c.fill = HEADER_FILL

        counts: dict[str, list[int]] = {}
        for r in rows:
            bucket = counts.setdefault(r[key], [0, 0, 0])
            bucket[0] += 1
            if not r["_ended"]:
                bucket[1] += 1
            if r["İlk Görülme"] == newest_date:
                bucket[2] += 1

        row_i = start_row + 1
        for name, (total, active, new) in sorted(counts.items(), key=lambda x: -x[1][0]):
            ws.cell(row=row_i, column=1, value=name).font = CELL_FONT
            ws.cell(row=row_i, column=2, value=total).font = CELL_FONT
            ws.cell(row=row_i, column=3, value=active).font = CELL_FONT
            c = ws.cell(row=row_i, column=4, value=new)
            c.font = CELL_FONT
            if new:
                c.fill = NEW_FILL
            row_i += 1
        return row_i + 2

    next_row = block(5, "Kategoriye Göre", "Kategori")
    block(next_row, "Bankaya Göre", "Banka")


def build_workbook(state_seen: dict, path: Path, use_llm: bool = True) -> tuple[Path, int, int]:
    """Excel'i sıfırdan üretir. (dosya_yolu, toplam_satır, bugün_eklenen) döner."""
    rows = build_rows(state_seen, use_llm=use_llm)
    newest_date = max((r["İlk Görülme"] for r in rows), default="")
    new_count = sum(1 for r in rows if r["İlk Görülme"] == newest_date)

    wb = Workbook()
    wb.remove(wb.active)

    ws_summary = wb.create_sheet("Özet")
    _write_summary(ws_summary, rows, newest_date)

    ws_all = wb.create_sheet("Tüm Kampanyalar")
    _write_sheet(ws_all, rows, newest_date, "TumKampanyalar")

    by_cat: dict[str, list[dict]] = {}
    for row in rows:
        by_cat.setdefault(row["Kategori"], []).append(row)

    for idx, (cat, cat_rows) in enumerate(
            sorted(by_cat.items(), key=lambda x: -len(x[1])), start=1):
        ws = wb.create_sheet(_safe_sheet_name(cat))
        _write_sheet(ws, cat_rows, newest_date, f"Kat{idx}")

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    log.info("Excel yazıldı: %s (%d satır, %d yeni, %d kategori)",
             path, len(rows), new_count, len(by_cat))
    return path, len(rows), new_count
